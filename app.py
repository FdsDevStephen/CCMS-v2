from __future__ import annotations

import gc
import html
import logging
import os
import time
import warnings

from datetime import datetime
from pathlib import Path

import streamlit as st
from openpyxl import Workbook, load_workbook


# ==========================================================
# QUIET LIBRARIES
# ==========================================================

os.environ["TOKENIZERS_PARALLELISM"] = "false"
os.environ["HF_HUB_DISABLE_TELEMETRY"] = "1"
os.environ["TRANSFORMERS_VERBOSITY"] = "error"
os.environ["HF_HUB_VERBOSITY"] = "error"

warnings.filterwarnings("ignore")

logging.getLogger("transformers").setLevel(
    logging.ERROR
)

logging.getLogger("huggingface_hub").setLevel(
    logging.ERROR
)

logging.getLogger("torch").setLevel(
    logging.ERROR
)


# ==========================================================
# PAGE CONFIG
# ==========================================================

st.set_page_config(
    page_title="CCMS",
    page_icon="⚖️",
    layout="wide",
    initial_sidebar_state="collapsed",
)


# ==========================================================
# IMPORTS
# ==========================================================

from ocr_v2 import OCRProcessor

from extractor.extractor import LegalExtractor

from extractor.prayer_extractor import PrayerExtractor

from extractor.survey_location import (
    SurveyLocationExtractor,
)

from RAG.chunker import LegalTextChunker

from RAG.embedding import EmbeddingModel

from RAG.retriever import LegalRetriever

from RAG.hybrid_retreiver import HybridRetriever

from RAG.vector_store import QdrantVectorStore

from RAG.act_extractor import ActExtractor


# ==========================================================
# PATHS
# ==========================================================

BASE_DIR = (
    Path(__file__)
    .resolve()
    .parent
)

OUTPUT_FOLDER = (
    BASE_DIR / "section_output"
)

OUTPUT_FOLDER.mkdir(
    parents=True,
    exist_ok=True,
)

RESULTS_FOLDER = (
    BASE_DIR / "results"
)

RESULTS_FOLDER.mkdir(
    parents=True,
    exist_ok=True,
)

UPLOAD_FOLDER = (
    BASE_DIR / ".streamlit_uploads"
)

UPLOAD_FOLDER.mkdir(
    parents=True,
    exist_ok=True,
)

EXCEL_PATH = (
    RESULTS_FOLDER
    / "legal_extraction_results.xlsx"
)

TESSERACT_PATH = (
    r"C:\Program Files\Tesseract-OCR\tesseract.exe"
)


# ==========================================================
# EXCEL COLUMNS
# ==========================================================

EXCEL_COLUMNS = [
    "Document",
    "Case Number",
    "Survey Numbers",
    "Survey Locations",
    "Acts",
    "Sections",
    "Act → Section Mapping",
    "Prayer",
    "Status",
    "Processing Time (s)",
    "Processed At",
]


# ==========================================================
# UI STYLING
# ==========================================================

st.markdown(
    """
    <style>

    .stApp {
        background: #0b0f14;
    }

    .block-container {
        max-width: 1180px;
        padding-top: 1.6rem;
        padding-bottom: 2.5rem;
    }

    header[data-testid="stHeader"] {
        background: #0b0f14;
    }

    h1,
    h2,
    h3 {
        color: #f8fafc !important;
        letter-spacing: -0.02em;
    }

    p,
    label,
    .stMarkdown {
        color: #cbd5e1;
    }

    [data-testid="stFileUploader"] {
        background: #11161d;
        border: 1px solid #27303d;
        border-radius: 12px;
        padding: 0.35rem;
    }

    [data-testid="stFileUploaderDropzone"] {
        background: #11161d;
        border-radius: 9px;
    }

    div.stButton > button {
        border-radius: 8px;
        min-height: 2.65rem;
        font-weight: 600;
        border: 1px solid #334155;
    }

    button[kind="primary"] {
        background: #2563eb !important;
        border-color: #2563eb !important;
    }

    button[kind="primary"]:hover {
        background: #1d4ed8 !important;
        border-color: #1d4ed8 !important;
    }

    [data-testid="stVerticalBlockBorderWrapper"] {
        background: #11161d;
        border: 1px solid #27303d;
        border-radius: 12px;
    }

    .result-title {
        font-size: 1.65rem;
        font-weight: 700;
        color: #f8fafc;
        margin-bottom: 0.15rem;
    }

    .result-subtitle {
        color: #94a3b8;
        font-size: 0.9rem;
        margin-bottom: 1.2rem;
    }

    .section-title {
        font-size: 1.15rem;
        font-weight: 700;
        color: #f1f5f9;
        margin: 0.25rem 0 0.7rem 0;
    }

    .prayer-box {
        background: #11161d;
        border: 1px solid #27303d;
        border-radius: 12px;
        padding: 1.2rem 1.35rem;
        line-height: 1.75;
        color: #e2e8f0;
        white-space: pre-wrap;
        font-size: 0.96rem;
    }

    .chip {
        display: inline-block;
        padding: 0.38rem 0.7rem;
        margin: 0.15rem 0.25rem 0.15rem 0;
        border-radius: 7px;
        background: #18202b;
        border: 1px solid #334155;
        color: #e2e8f0;
        font-size: 0.88rem;
    }

    .mapping-act {
        color: #f8fafc;
        font-weight: 650;
        margin-bottom: 0.7rem;
    }

    .location-row {
        padding: 0.7rem 0;
        border-bottom: 1px solid #222b36;
    }

    .location-row:last-child {
        border-bottom: none;
    }

    .location-survey {
        color: #f8fafc;
        font-weight: 650;
        margin-bottom: 0.25rem;
    }

    .location-details {
        color: #94a3b8;
        font-size: 0.9rem;
    }

    .upload-heading {
        font-size: 1.05rem;
        font-weight: 650;
        color: #f8fafc;
        margin-bottom: 0.2rem;
    }

    </style>
    """,
    unsafe_allow_html=True,
)


# ==========================================================
# EXCEL HELPERS
# ==========================================================

def ensure_excel_file() -> None:

    if EXCEL_PATH.exists():
        return

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Results"

    worksheet.append(
        EXCEL_COLUMNS
    )

    worksheet.freeze_panes = "A2"

    workbook.save(
        EXCEL_PATH
    )

    workbook.close()


def save_excel_workbook(
    workbook,
) -> bool:

    for attempt in range(3):

        try:

            workbook.save(
                EXCEL_PATH
            )

            return True

        except (
            PermissionError,
            OSError,
        ):

            if attempt < 2:

                time.sleep(1)

            else:

                return False

    return False


def get_act_name(
    act,
) -> str:

    if isinstance(
        act,
        dict,
    ):

        return str(
            act.get(
                "name",
                "",
            )
        ).strip()

    return str(
        act
    ).strip()


def format_survey_locations(
    locations,
) -> str:

    if not isinstance(
        locations,
        list,
    ):

        return ""

    lines = []

    for location in locations:

        if not isinstance(
            location,
            dict,
        ):

            continue

        survey = str(
            location.get(
                "survey_number",
                "",
            )
            or ""
        ).strip()

        parts = []

        village = location.get(
            "village"
        )

        hobli = location.get(
            "hobli"
        )

        taluk = location.get(
            "taluk"
        )

        district = location.get(
            "district"
        )

        if village:
            parts.append(
                f"Village: {village}"
            )

        if hobli:
            parts.append(
                f"Hobli: {hobli}"
            )

        if taluk:
            parts.append(
                f"Taluk: {taluk}"
            )

        if district:
            parts.append(
                f"District: {district}"
            )

        if parts:

            if survey:

                lines.append(
                    f"{survey}: "
                    + ", ".join(parts)
                )

            else:

                lines.append(
                    ", ".join(parts)
                )

    return "\n".join(
        lines
    )


def format_mapping(
    mappings,
) -> str:

    if not isinstance(
        mappings,
        list,
    ):

        return ""

    lines = []

    for mapping in mappings:

        if not isinstance(
            mapping,
            dict,
        ):

            continue

        act = str(
            mapping.get(
                "act",
                "",
            )
            or ""
        ).strip()

        sections = mapping.get(
            "sections",
            [],
        )

        if sections is None:

            sections = []

        elif not isinstance(
            sections,
            list,
        ):

            sections = [
                sections
            ]

        sections = [
            str(section).strip()
            for section in sections
            if section is not None
            and str(section).strip()
        ]

        if not act:
            continue

        if sections:

            lines.append(
                f"{act}: "
                + ", ".join(
                    sections
                )
            )

        else:

            lines.append(
                f"{act}:"
            )

    return "\n".join(
        lines
    )


def prepare_excel_values(
    result: dict,
) -> dict:

    survey_numbers = result.get(
        "survey_numbers",
        [],
    )

    acts = result.get(
        "acts",
        [],
    )

    sections = result.get(
        "sections",
        [],
    )

    locations = result.get(
        "survey_locations",
        [],
    )

    mappings = result.get(
        "act_section_mapping",
        [],
    )

    prayer = result.get(
        "prayer",
        "",
    )

    survey_numbers_text = (
        ", ".join(
            str(value)
            for value in survey_numbers
        )
    )

    acts_text = "\n".join(
        get_act_name(act)
        for act in acts
        if get_act_name(act)
    )

    sections_text = (
        ", ".join(
            str(section)
            for section in sections
        )
    )

    return {
        "survey_numbers":
            survey_numbers_text,

        "survey_locations":
            format_survey_locations(
                locations
            ),

        "acts":
            acts_text,

        "sections":
            sections_text,

        "mapping":
            format_mapping(
                mappings
            ),

        "prayer":
            str(
                prayer
                or ""
            ),
    }


def append_result_to_excel(
    result: dict,
    document_name: str,
    processing_time: float,
    status: str,
) -> bool:

    ensure_excel_file()

    workbook = load_workbook(
        EXCEL_PATH
    )

    worksheet = workbook[
        "Results"
    ]

    values = prepare_excel_values(
        result
    )

    worksheet.append(
        [
            document_name,

            result.get(
                "case_number",
                "",
            ),

            values[
                "survey_numbers"
            ],

            values[
                "survey_locations"
            ],

            values[
                "acts"
            ],

            values[
                "sections"
            ],

            values[
                "mapping"
            ],

            values[
                "prayer"
            ],

            status,

            round(
                processing_time,
                2,
            ),

            datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            ),
        ]
    )

    worksheet.freeze_panes = "A2"

    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

    for column in worksheet.columns:

        maximum_length = 0

        column_letter = (
            column[0].column_letter
        )

        for cell in column:

            try:

                maximum_length = max(
                    maximum_length,
                    len(
                        str(
                            cell.value
                        )
                    ),
                )

            except Exception:
                pass

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max(
                maximum_length + 2,
                12,
            ),
            60,
        )

    for row in worksheet.iter_rows():

        for cell in row:

            cell.alignment = (
                cell.alignment.copy(
                    wrap_text=True,
                    vertical="top",
                )
            )

    saved = save_excel_workbook(
        workbook
    )

    workbook.close()

    return saved


# ==========================================================
# MEMORY CLEANUP
# ==========================================================

def cleanup_memory() -> None:

    gc.collect()

    try:

        import torch

        if torch.cuda.is_available():
            torch.cuda.empty_cache()

    except Exception:
        pass


# ==========================================================
# CACHED RESOURCES
# ==========================================================

@st.cache_resource(
    show_spinner="Loading BGE-M3..."
)
def get_embedding_model():

    return EmbeddingModel()


@st.cache_resource
def get_ocr_processor():

    return OCRProcessor(
        output_folder=OUTPUT_FOLDER,
        tesseract_path=TESSERACT_PATH,
        search_pages=30,
        fast_dpi=150,
        full_dpi=220,

        # IMPORTANT:
        # Keep this low on an 8 GB machine.
        max_workers=1,

        page_start=2,
        page_end=13,
        prefer_text_layer=True,
        denoise=True,
    )


@st.cache_resource
def get_legal_extractor():

    return LegalExtractor()


# ==========================================================
# UPLOAD HELPER
# ==========================================================

def save_uploaded_pdf(
    pdf_bytes: bytes,
    filename: str,
) -> Path:

    safe_name = Path(
        filename
    ).name

    destination = (
        UPLOAD_FOLDER
        / safe_name
    )

    destination.write_bytes(
        pdf_bytes
    )

    return destination


# ==========================================================
# MAIN PIPELINE
# ==========================================================

def run_pipeline(
    pdf_bytes: bytes,
    filename: str,
):

    pipeline_start = (
        time.perf_counter()
    )

    timings = {}

    document_name = (
        Path(filename).stem
    )

    # ======================================================
    # SAVE PDF
    # ======================================================

    pdf_path = save_uploaded_pdf(
        pdf_bytes,
        filename,
    )

    # ======================================================
    # 1. OCR
    # ======================================================

    stage_start = (
        time.perf_counter()
    )

    ocr_processor = (
        get_ocr_processor()
    )

    ocr_result = (
        ocr_processor.process(
            pdf_path
        )
    )

    if isinstance(
        ocr_result,
        tuple,
    ):

        ocr_text = (
            ocr_result[0]
        )

        txt_path = (
            ocr_result[1]
        )

    else:

        ocr_text = (
            ocr_result
        )

        txt_path = (
            OUTPUT_FOLDER
            / f"{document_name}.txt"
        )

    if not isinstance(
        ocr_text,
        str,
    ):

        raise RuntimeError(
            "OCR did not return text."
        )

    if not txt_path.exists():

        txt_path = (
            OUTPUT_FOLDER
            / f"{document_name}.txt"
        )

    if not txt_path.exists():

        txt_path.write_text(
            ocr_text,
            encoding="utf-8",
        )

    timings[
        "OCR"
    ] = (
        time.perf_counter()
        - stage_start
    )

    # ======================================================
    # 2. LEGAL EXTRACTION
    # ======================================================

    stage_start = (
        time.perf_counter()
    )

    legal_extractor = (
        get_legal_extractor()
    )

    base_result = (
        legal_extractor.extract(
            text=ocr_text,
            case_number=document_name,
        )
    )

    timings[
        "Legal Extraction"
    ] = (
        time.perf_counter()
        - stage_start
    )

    # ======================================================
    # 3. PRAYER
    # ======================================================

    stage_start = (
        time.perf_counter()
    )

    prayer_extractor = (
        PrayerExtractor()
    )

    prayer = (
        prayer_extractor.extract(
            ocr_text
        )
    )

    base_result[
        "prayer"
    ] = prayer

    timings[
        "Prayer"
    ] = (
        time.perf_counter()
        - stage_start
    )

    # ======================================================
    # 4. BGE-M3
    # ======================================================

    stage_start = (
        time.perf_counter()
    )

    embedding_model = (
        get_embedding_model()
    )

    timings[
        "BGE-M3"
    ] = (
        time.perf_counter()
        - stage_start
    )

    # ======================================================
    # 5. CHUNKING
    # ======================================================

    stage_start = (
        time.perf_counter()
    )

    chunker = LegalTextChunker(
        chunk_size=450,
        overlap=50,
        tokenizer=(
            embedding_model.tokenizer
        ),
    )

    chunks = (
        chunker.chunk_file(
            txt_path
        )
    )

    if not chunks:

        raise RuntimeError(
            "No chunks were created."
        )

    timings[
        "Chunking"
    ] = (
        time.perf_counter()
        - stage_start
    )

    # ======================================================
    # 6. EMBEDDINGS
    # ======================================================

    stage_start = (
        time.perf_counter()
    )

    texts = [
        chunk["text"]
        for chunk in chunks
    ]

    embeddings = (
        embedding_model.encode(
            texts,
            batch_size=12,
        )
    )

    timings[
        "Embeddings"
    ] = (
        time.perf_counter()
        - stage_start
    )

    # ======================================================
    # 7. QDRANT
    # ======================================================

    stage_start = (
        time.perf_counter()
    )

    vector_store = (
        QdrantVectorStore()
    )

    vector_store.insert(
        chunks,
        embeddings,
    )

    timings[
        "Qdrant"
    ] = (
        time.perf_counter()
        - stage_start
    )

    # ======================================================
    # 8. HYBRID RETRIEVER
    # ======================================================

    stage_start = (
        time.perf_counter()
    )

    legal_retriever = (
        LegalRetriever(
            embedding_model=embedding_model,
        )
    )

    hybrid_retriever = (
        HybridRetriever(
            vector_retriever=legal_retriever,
        )
    )

    hybrid_retriever.build_bm25(
        chunks
    )

    timings[
        "BM25 + Hybrid"
    ] = (
        time.perf_counter()
        - stage_start
    )

    # ======================================================
    # 9. ACT + SECTION EXTRACTION
    # ======================================================

    stage_start = (
        time.perf_counter()
    )

    act_extractor = (
        ActExtractor(
            chunks,
            retriever=hybrid_retriever,
        )
    )

    act_result = (
        act_extractor.extract(
            document=document_name,
            sections=base_result.get(
                "sections",
                [],
            ),
            top_k=5,
        )
    )

    timings[
        "Act Extraction"
    ] = (
        time.perf_counter()
        - stage_start
    )

    # ======================================================
    # 10. SURVEY LOCATIONS
    # ======================================================

    stage_start = (
        time.perf_counter()
    )

    survey_numbers = (
        base_result.get(
            "survey_numbers",
            [],
        )
    )

    if survey_numbers:

        location_extractor = (
            SurveyLocationExtractor(
                ocr_text
            )
        )

        survey_locations = (
            location_extractor.extract(
                survey_numbers
            )
        )

    else:

        survey_locations = []

    base_result[
        "survey_locations"
    ] = survey_locations

    timings[
        "Survey Locations"
    ] = (
        time.perf_counter()
        - stage_start
    )

    # ======================================================
    # 11. FINAL RESULT
    # ======================================================

    base_result[
        "acts"
    ] = act_result.get(
        "acts",
        [],
    )

    base_result[
        "sections"
    ] = act_result.get(
        "sections",
        base_result.get(
            "sections",
            [],
        ),
    )

    base_result[
        "act_section_mapping"
    ] = act_result.get(
        "act_section_mapping",
        [],
    )

    base_result[
        "primary_act"
    ] = act_result.get(
        "primary_act",
        base_result.get(
            "primary_act",
            None,
        ),
    )

    timings[
        "Total"
    ] = (
        time.perf_counter()
        - pipeline_start
    )

    return (
        base_result,
        timings,
        ocr_text,
        chunks,
    )


# ==========================================================
# SESSION STATE
# ==========================================================

if "result" not in st.session_state:
    st.session_state.result = None

if "timings" not in st.session_state:
    st.session_state.timings = None

if "ocr_text" not in st.session_state:
    st.session_state.ocr_text = None

if "chunks" not in st.session_state:
    st.session_state.chunks = None

if "filename" not in st.session_state:
    st.session_state.filename = None


# ==========================================================
# HEADER
# ==========================================================

st.title(
    "⚖️ CCMS"
)

st.caption(
    "Karnataka High Court Legal Document Analysis"
)

st.divider()


# ==========================================================
# UPLOAD
# ==========================================================

st.markdown(
    '<div class="upload-heading">'
    'Upload Document'
    '</div>',
    unsafe_allow_html=True,
)

upload_mode = st.radio(
    "Input",
    [
        "📄 Single PDF",
        "📁 Folder",
    ],
    horizontal=True,
    label_visibility="collapsed",
    key="upload_mode",
)

uploaded_file = None
folder_files = []


if upload_mode == "📄 Single PDF":

    uploaded_file = st.file_uploader(
        "Upload a Karnataka High Court PDF",
        type=["pdf"],
        key="single_pdf",
    )

else:

    folder_files = st.file_uploader(
        "Upload a folder containing PDFs",
        type=["pdf"],
        accept_multiple_files="directory",
        key="folder_pdf",
    )


# ==========================================================
# FILE SELECTION
# ==========================================================

if upload_mode == "📁 Folder":

    files_to_process = [
        file
        for file in folder_files
        if file.name.lower().endswith(
            ".pdf"
        )
    ]

else:

    files_to_process = (
        [uploaded_file]
        if uploaded_file is not None
        else []
    )


if files_to_process:

    st.caption(
        f"{len(files_to_process)} PDF"
        f"{'' if len(files_to_process) == 1 else 's'} "
        "selected."
    )

    process_clicked = st.button(
        "⚡ Analyze Document"
        if len(files_to_process) == 1
        else "⚡ Analyze Documents",
        type="primary",
        use_container_width=True,
        key="analyze_documents",
    )

else:

    process_clicked = False


# ==========================================================
# PROCESS
# ==========================================================

if process_clicked:

    st.session_state.result = None
    st.session_state.timings = None
    st.session_state.ocr_text = None
    st.session_state.chunks = None
    st.session_state.filename = None

    ensure_excel_file()

    total_files = (
        len(files_to_process)
    )

    successful = 0
    failed = 0

    progress = st.progress(
        0,
        text="Preparing...",
    )

    status_box = st.empty()

    for index, uploaded in enumerate(
        files_to_process,
        start=1,
    ):

        filename = Path(
            uploaded.name
        ).name

        status_box.info(
            f"Analyzing {filename}  •  "
            f"{index}/{total_files}"
        )

        document_start = (
            time.perf_counter()
        )

        pdf_bytes = None
        result = None
        timings = None
        ocr_text = None
        chunks = None

        try:

            # Read only the current PDF.
            pdf_bytes = (
                uploaded.getvalue()
            )

            (
                result,
                timings,
                ocr_text,
                chunks,
            ) = run_pipeline(
                pdf_bytes=pdf_bytes,
                filename=filename,
            )

            document_time = (
                time.perf_counter()
                - document_start
            )

            excel_saved = (
                append_result_to_excel(
                    result=result,
                    document_name=filename,
                    processing_time=document_time,
                    status="Success",
                )
            )

            successful += 1

            # Keep only the latest result visible.
            st.session_state.result = (
                result
            )

            st.session_state.timings = (
                timings
            )

            st.session_state.ocr_text = (
                ocr_text
            )

            st.session_state.chunks = (
                chunks
            )

            st.session_state.filename = (
                filename
            )

            if not excel_saved:

                st.warning(
                    "Document analysis completed, "
                    "but the Excel file could not be "
                    "updated. Please close "
                    "legal_extraction_results.xlsx "
                    "if it is open."
                )

        except Exception as exc:

            failed += 1

            document_time = (
                time.perf_counter()
                - document_start
            )

            failed_result = {
                "case_number":
                    Path(
                        filename
                    ).stem,

                "survey_numbers":
                    [],

                "survey_locations":
                    [],

                "acts":
                    [],

                "sections":
                    [],

                "act_section_mapping":
                    [],

                "primary_act":
                    None,

                "prayer":
                    "",
            }

            excel_saved = (
                append_result_to_excel(
                    result=failed_result,
                    document_name=filename,
                    processing_time=document_time,
                    status=f"Failed: {exc}",
                )
            )

            if not excel_saved:

                st.warning(
                    "The Excel file could not be "
                    "updated because it is locked. "
                    "Close "
                    "legal_extraction_results.xlsx "
                    "and try again."
                )

            st.error(
                f"Failed to process "
                f"{filename}: {exc}"
            )

        finally:

            # Release the uploaded PDF immediately.
            if pdf_bytes is not None:

                del pdf_bytes

            # IMPORTANT:
            # This prevents memory from accumulating
            # while processing a folder.
            cleanup_memory()

        progress.progress(
            int(
                index
                / total_files
                * 100
            ),
            text=(
                f"Processed "
                f"{index}/{total_files}"
            ),
        )

    status_box.empty()

    progress.empty()

    if successful:

        st.success(
            f"{successful} document"
            f"{'' if successful == 1 else 's'} "
            "processed successfully."
        )

    if failed:

        st.error(
            f"{failed} document"
            f"{'' if failed == 1 else 's'} "
            "failed."
        )


# ==========================================================
# RESULT
# ==========================================================

result = (
    st.session_state.result
)


if result is None:

    st.divider()

    with st.container(
        border=True
    ):

        st.markdown(
            '<div class="result-title">'
            'Analyze a Legal Document'
            '</div>',
            unsafe_allow_html=True,
        )

        st.caption(
            "Upload a PDF above to extract "
            "case information, survey details, "
            "Acts, Sections and Prayer."
        )

    st.stop()


# ==========================================================
# RESULT HEADER
# ==========================================================

st.divider()

filename = (
    st.session_state.filename
    or "Legal Document"
)

case_number = result.get(
    "case_number",
    "N/A",
)

st.markdown(
    '<div class="result-title">'
    'Analysis Result'
    '</div>',
    unsafe_allow_html=True,
)

st.markdown(
    f'<div class="result-subtitle">'
    f'{html.escape(filename)}'
    f' &nbsp;•&nbsp; '
    f'Case No. '
    f'{html.escape(str(case_number))}'
    f'</div>',
    unsafe_allow_html=True,
)


# ==========================================================
# ACTS + SECTIONS
# ==========================================================

left, right = st.columns(
    [1, 1],
    gap="large",
)


with left:

    st.markdown(
        '<div class="section-title">'
        '📜 Acts'
        '</div>',
        unsafe_allow_html=True,
    )

    acts = result.get(
        "acts",
        [],
    )

    if acts:

        for act in acts:

            act_name = get_act_name(
                act
            )

            if act_name:

                st.markdown(
                    f'<div class="chip">'
                    f'📜 '
                    f'{html.escape(act_name)}'
                    f'</div>',
                    unsafe_allow_html=True,
                )

    else:

        st.caption(
            "No Acts found."
        )


with right:

    st.markdown(
        '<div class="section-title">'
        '§ Sections'
        '</div>',
        unsafe_allow_html=True,
    )

    sections = result.get(
        "sections",
        [],
    )

    if sections:

        for section in sections:

            st.markdown(
                f'<span class="chip">'
                f'§ '
                f'{html.escape(str(section))}'
                f'</span>',
                unsafe_allow_html=True,
            )

    else:

        st.caption(
            "No Sections found."
        )


st.write("")


# ==========================================================
# SURVEY NUMBERS + LOCATIONS
# ==========================================================

left, right = st.columns(
    [1, 1],
    gap="large",
)


with left:

    st.markdown(
        '<div class="section-title">'
        '📍 Survey Numbers'
        '</div>',
        unsafe_allow_html=True,
    )

    survey_numbers = result.get(
        "survey_numbers",
        [],
    )

    if survey_numbers:

        for survey in survey_numbers:

            st.markdown(
                f'<span class="chip">'
                f'{html.escape(str(survey))}'
                f'</span>',
                unsafe_allow_html=True,
            )

    else:

        st.caption(
            "No Survey Numbers found."
        )


with right:

    st.markdown(
        '<div class="section-title">'
        '🗺️ Survey Locations'
        '</div>',
        unsafe_allow_html=True,
    )

    survey_locations = result.get(
        "survey_locations",
        [],
    )

    if survey_locations:

        with st.container(
            border=True
        ):

            for location in (
                survey_locations
            ):

                if not isinstance(
                    location,
                    dict,
                ):

                    continue

                survey = str(
                    location.get(
                        "survey_number",
                        "",
                    )
                    or ""
                ).strip()

                parts = []

                for key, label in [
                    (
                        "village",
                        "Village",
                    ),
                    (
                        "hobli",
                        "Hobli",
                    ),
                    (
                        "taluk",
                        "Taluk",
                    ),
                    (
                        "district",
                        "District",
                    ),
                ]:

                    value = (
                        location.get(
                            key
                        )
                    )

                    if value:

                        parts.append(
                            f"{label}: "
                            f"{value}"
                        )

                st.markdown(
                    '<div class="location-row">'
                    '<div class="location-survey">'
                    f'{html.escape(survey)}'
                    '</div>'
                    '<div class="location-details">'
                    f'{html.escape(", ".join(parts))}'
                    '</div>'
                    '</div>',
                    unsafe_allow_html=True,
                )

    else:

        st.caption(
            "No Survey Locations found."
        )


st.write("")


# ==========================================================
# ACT → SECTION MAPPING
# ==========================================================

st.markdown(
    '<div class="section-title">'
    '🔗 Act → Section Mapping'
    '</div>',
    unsafe_allow_html=True,
)

mappings = result.get(
    "act_section_mapping",
    [],
)

valid_mappings = []


for mapping in mappings:

    if not isinstance(
        mapping,
        dict,
    ):

        continue

    act = str(
        mapping.get(
            "act",
            "",
        )
        or ""
    ).strip()

    mapped_sections = (
        mapping.get(
            "sections",
            [],
        )
    )

    if mapped_sections is None:

        mapped_sections = []

    elif not isinstance(
        mapped_sections,
        list,
    ):

        mapped_sections = [
            mapped_sections
        ]

    mapped_sections = [
        str(section).strip()
        for section in mapped_sections
        if section is not None
        and str(section).strip()
    ]

    if act:

        valid_mappings.append(
            (
                act,
                mapped_sections,
            )
        )


if valid_mappings:

    for (
        act,
        mapped_sections,
    ) in valid_mappings:

        with st.container(
            border=True
        ):

            st.markdown(
                f'<div class="mapping-act">'
                f'📜 '
                f'{html.escape(act)}'
                f'</div>',
                unsafe_allow_html=True,
            )

            if mapped_sections:

                for section in (
                    mapped_sections
                ):

                    st.markdown(
                        f'<span class="chip">'
                        f'§ '
                        f'{html.escape(section)}'
                        f'</span>',
                        unsafe_allow_html=True,
                    )

            else:

                st.caption(
                    "No sections mapped."
                )

else:

    st.caption(
        "No Act → Section mapping found."
    )


# ==========================================================
# PRAYER — ALWAYS AT THE END
# ==========================================================

st.divider()

st.markdown(
    '<div class="section-title">'
    '🙏 Prayer'
    '</div>',
    unsafe_allow_html=True,
)

prayer = result.get(
    "prayer",
    "",
)

if prayer:

    st.markdown(
        '<div class="prayer-box">'
        + html.escape(
            str(prayer)
        )
        + '</div>',
        unsafe_allow_html=True,
    )

else:

    st.caption(
        "No Prayer found."
    )


# ==========================================================
# FOOTER
# ==========================================================

st.divider()

st.caption(
    "CCMS · Karnataka High Court "
    "Legal Document Analysis"
)